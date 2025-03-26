from io import BytesIO
import pandas as pd
from utils.db import db
from models.solicitudes import Solicitud
from models.facturas import Factura
from models.proveedores_calificados import ProveedorCalificado
from models.parametros import Parametro
from models.desembolsos import Desembolso
from flask import jsonify, send_file
import shutil
from datetime import datetime
from openpyxl import load_workbook
import os
from jinja2 import Environment, FileSystemLoader
import pdfkit
from io import BytesIO

def actualizar_solicitudes(lista_ids):
    """
    Actualiza el estado de solicitudes a 'Procesada' (id_estado = 4)
    y crea registros en la tabla desembolsos para cada solicitud actualizada.
    """
    if not lista_ids or not isinstance(lista_ids, list):
        return None, "Debe proporcionar una lista de valores únicos válida"

    solicitudes_a_procesar = Solicitud.query.filter(
        Solicitud.id.in_(lista_ids),
        Solicitud.id_estado != 4
    ).all()

    if not solicitudes_a_procesar:
        return None, "No se encontraron solicitudes para actualizar"

    for solicitud in solicitudes_a_procesar:
        solicitud.id_estado = 4

        cuenta_bancaria = None
        if solicitud.factura and solicitud.factura.proveedor:
            cuenta_bancaria = solicitud.factura.proveedor.cuenta_bancaria

        nuevo_desembolso = Desembolso(
            fecha_desembolso=datetime.now().strftime('%Y/%m/%d'),
            monto_final=solicitud.total,
            metodo_pago='Transferencia bancaria',
            cuenta_bancaria=cuenta_bancaria,
            no_transaccion=None,
            estado=5,
            id_solicitud=solicitud.id,
            created_at=datetime.now(),
            updated_at=datetime.now(),
            descripcion=f"Desembolso factoraje #{solicitud.id}"
        )

        db.session.add(nuevo_desembolso)

    db.session.commit()

    return f'Se procesaron: {len(solicitudes_a_procesar)} solicitudes.', None


def exportar_solicitudes(lista_ids, formato="excel", ruta_archivo_base="utils/plantilla_desembolsos.xlsx", numero_inicial=1):
    """
    Extrae la información de las solicitudes, copia un archivo base de Excel,
    inserta los datos en celdas específicas y devuelve el archivo generado.
    """
    ruta_archivo_base_absoluta = os.path.abspath(ruta_archivo_base)

    # Verificar si el archivo base existe
    if not os.path.exists(ruta_archivo_base_absoluta):
        return jsonify({"error": f"El archivo base no existe en la ruta: {ruta_archivo_base_absoluta}"}), 400

    # Imprimir la ruta absoluta para verificar
    print(f"Ruta absoluta del archivo base: {ruta_archivo_base_absoluta}")

    # Obtener el valor de la empresa desde la tabla parámetros
    empresa_parametro = db.session.query(Parametro.valor).filter(Parametro.clave == "NOM-EMPRESA").first()
    empresa = empresa_parametro.valor if empresa_parametro else "Desconocido"

    # Consulta a la base de datos
    query = db.session.query(
        Solicitud.id,
        Solicitud.total,
        Factura.no_factura.label("numero_factura"),
        Factura.monto.label("monto_factura"),
        Factura.fecha_emision.label("fecha_factura"),
        ProveedorCalificado.razon_social.label("nombre_proveedor"),
        ProveedorCalificado.cuenta_bancaria.label("numero_cuenta_bancaria_proveedor"),
        ProveedorCalificado.correo_electronico.label("email_proveedor"),
        ProveedorCalificado.codigo_banco,
        ProveedorCalificado.nit.label("nit_proveedor"),
    ).join(Factura, Solicitud.id_factura == Factura.id)\
     .join(ProveedorCalificado, Factura.id_proveedor == ProveedorCalificado.id)\
     .filter(Solicitud.id.in_(lista_ids))

    resultados = query.all()

    fecha_actual = datetime.now()
    año = fecha_actual.year
    mes = str(fecha_actual.month).zfill(2)
    dia = str(fecha_actual.day).zfill(2)

    # Definir el mapeo de columnas
    mapeo_columnas = {
        "numero_cuenta_bancaria_proveedor": 4,  # Columna D
        "numero_registro": 5,                   # Columna E - AUTOINCREMENTAL
        "año": 6,                               # Columna F
        "mes": 7,                               # Columna G
        "dia": 8,                               # Columna H
        "total": 9,                             # Columna I
        "suma_registros": 10,                   # Columna J
        "empresa": 11,                          # Columna K
        "nombre_proveedor": 13,                 # Columna M
        "codigo_banco": 15,                     # Columna O
        "descripcion": 16,                      # Columna P
        "nit_proveedor": 17,                    # Columna Q
    }

    suma_totales = sum(int(str(r.total).replace('.', '')) for r in resultados)
    cantidad_registros = len(resultados)

    # Insertar información global en la primera fila
    data = [
        [
            r.numero_cuenta_bancaria_proveedor,
            numero_inicial + index,  # Número autoincremental
            año,
            mes,
            dia,
            int(str(r.total).replace('.', '')),
            cantidad_registros if index == 0 else numero_inicial + index,  # Cantidad de registros en la primera fila, número incremental en las siguientes
            empresa,
            r.nombre_proveedor,
            r.codigo_banco,
            f"Desembolso factoraje #{r.id}",  # Descripción única por solicitud
            r.nit_proveedor,
        ]
        for index, r in enumerate(resultados)
    ]

    fecha_actual_str = fecha_actual.strftime("%d-%m-%Y")
    nombre_archivo = f"Desembolsos_{fecha_actual_str}.xlsx"

    # Copiar el archivo base
    shutil.copy2(ruta_archivo_base_absoluta, nombre_archivo)

    libro = load_workbook(nombre_archivo)
    hoja = libro.active

    # Colocar datos globales en la primera fila
    hoja.cell(row=1, column=mapeo_columnas["año"], value=año)
    hoja.cell(row=1, column=mapeo_columnas["mes"], value=mes)
    hoja.cell(row=1, column=mapeo_columnas["dia"], value=dia)
    hoja.cell(row=1, column=mapeo_columnas["total"], value=suma_totales)
    hoja.cell(row=1, column=mapeo_columnas["suma_registros"], value=cantidad_registros)

    # Insertar los datos en Excel desde la segunda fila
    for i, fila in enumerate(data, start=2):
        for clave, valor in zip(mapeo_columnas.keys(), fila):
            columna = mapeo_columnas[clave]
            if clave == "suma_registros":
                hoja.cell(row=i, column=columna, value=i-1)  # Valor incremental a partir de 1
            else:
                hoja.cell(row=i, column=columna, value=valor)

    output = BytesIO()
    libro.save(output)
    libro.close()
    output.seek(0)

    # Eliminar el archivo temporal
    os.remove(nombre_archivo)

    return send_file(
        output,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
        

def generar_pdf_solicitud(solicitud):
    """
    Genera un PDF con los detalles de la solicitud usando una plantilla HTML.
    
    Args:
        solicitud: Objeto de solicitud con los datos necesarios
        
    Returns:
        BytesIO: Buffer con el contenido del PDF generado
    """
    try:
        # Configurar el entorno de Jinja2
        template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates')
        env = Environment(loader=FileSystemLoader(template_dir))
        template = env.get_template('solicitud_pdf.html')

        # Preparar los datos para la plantilla
        datos_solicitud = {
            'numero_factura': solicitud.factura.no_factura,
            'fecha_otorgamiento': solicitud.factura.fecha_otorga.strftime('%d/%m/%Y'),
            'fecha_vencimiento': solicitud.factura.fecha_vence.strftime('%d/%m/%Y'),
            'monto_factura': float(solicitud.factura.monto),
            'descuento': float(solicitud.descuento_app or 0),
            'iva': float(solicitud.iva or 0),
            'subtotal_descuento': float(solicitud.subtotal or 0),
            'total_recibir': float(solicitud.total or 0),
            'nombre_proveedor': solicitud.factura.nombre_proveedor,
            'nit_proveedor': solicitud.factura.nit,
            'nombre_cliente': solicitud.factura.proveedor.razon_social if solicitud.factura and solicitud.factura.proveedor else 'No especificado'
        }

        # Renderizar la plantilla
        html_content = template.render(
            solicitud=datos_solicitud,
            fecha_generacion=datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        )

        # Configuración de wkhtmltopdf
        config = pdfkit.configuration(wkhtmltopdf='C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe')
        
        # Opciones para el PDF
        options = {
            'page-size': 'Letter',
            'margin-top': '1.85in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'encoding': "UTF-8",
            'no-outline': None
        }

        # Generar PDF
        pdf = pdfkit.from_string(html_content, False, options=options, configuration=config)
        
        # Crear buffer de memoria con el PDF
        pdf_buffer = BytesIO(pdf)
        pdf_buffer.seek(0)
        
        return pdf_buffer
    except Exception as e:
        raise Exception(f"Error al generar el PDF: {str(e)}")
